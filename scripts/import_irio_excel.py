"""
ETL script: Extract IRIO data from Excel workbook into processed binary/JSON files.

Reads IRIO_Analysis_Template_8.xlsx and produces:
- leontief_inverse.npy (578x578 float64 matrix)
- province_sector_data.json (output, VA, VA coefficient per province-sector)
- province_metadata.json (province list with codes, names, regions)
- sector_metadata.json (sector list with codes and names)
- province_pdrb.json (PDRB summary per province)
- bilateral_linkage.npy (34x34 bilateral linkage matrix)
- bilateral_labels.json (row/col labels for bilateral matrix)
- fl_bl_data.json (FL/BL per province-sector from formulas)
- region_mapping.json (province-to-region mapping)
"""

import json
import logging
import sys
from pathlib import Path

import numpy as np
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
OUT_DIR = PROJECT_ROOT / "data" / "processed"
WORKBOOK_PATH = RAW_DIR / "IRIO_Analysis_Template_8.xlsx"

PROVINCE_ORDER = [
    ("11", "Aceh", "Sumatera"),
    ("12", "Sumatera Utara", "Sumatera"),
    ("13", "Sumatera Barat", "Sumatera"),
    ("14", "Riau", "Sumatera"),
    ("15", "Jambi", "Sumatera"),
    ("16", "Sumatera Selatan", "Sumatera"),
    ("17", "Bengkulu", "Sumatera"),
    ("18", "Lampung", "Sumatera"),
    ("19", "Kep. Bangka Belitung", "Sumatera"),
    ("21", "Kep. Riau", "Sumatera"),
    ("31", "DKI Jakarta", "Jawa"),
    ("32", "Jawa Barat", "Jawa"),
    ("33", "Jawa Tengah", "Jawa"),
    ("34", "DI Yogyakarta", "Jawa"),
    ("35", "Jawa Timur", "Jawa"),
    ("36", "Banten", "Jawa"),
    ("51", "Bali", "Bali-Nusa Tenggara"),
    ("52", "Nusa Tenggara Barat", "Bali-Nusa Tenggara"),
    ("53", "Nusa Tenggara Timur", "Bali-Nusa Tenggara"),
    ("61", "Kalimantan Barat", "Kalimantan"),
    ("62", "Kalimantan Tengah", "Kalimantan"),
    ("63", "Kalimantan Selatan", "Kalimantan"),
    ("64", "Kalimantan Timur", "Kalimantan"),
    ("65", "Kalimantan Utara", "Kalimantan"),
    ("71", "Sulawesi Utara", "Sulawesi"),
    ("72", "Sulawesi Tengah", "Sulawesi"),
    ("73", "Sulawesi Selatan", "Sulawesi"),
    ("74", "Sulawesi Tenggara", "Sulawesi"),
    ("75", "Gorontalo", "Sulawesi"),
    ("76", "Sulawesi Barat", "Sulawesi"),
    ("81", "Maluku", "Maluku-Papua"),
    ("82", "Maluku Utara", "Maluku-Papua"),
    ("91", "Papua Barat", "Maluku-Papua"),
    ("94", "Papua", "Maluku-Papua"),
]

SECTOR_ORDER = [
    ("A", "Pertanian, Kehutanan, dan Perikanan"),
    ("B", "Pertambangan dan Penggalian"),
    ("C", "Industri Pengolahan"),
    ("D", "Pengadaan Listrik dan Gas"),
    ("E", "Pengadaan Air, Pengelolaan Sampah, Limbah dan Daur Ulang"),
    ("F", "Konstruksi"),
    ("G", "Perdagangan Besar dan Eceran; Reparasi Mobil dan Sepeda Motor"),
    ("H", "Transportasi dan Pergudangan"),
    ("I", "Penyediaan Akomodasi dan Makan Minum"),
    ("J", "Informasi dan Komunikasi"),
    ("K", "Jasa Keuangan dan Asuransi"),
    ("L", "Real Estate"),
    ("MN", "Jasa Perusahaan"),
    ("O", "Administrasi Pemerintahan, Pertahanan dan Jaminan Sosial Wajib"),
    ("P", "Jasa Pendidikan"),
    ("Q", "Jasa Kesehatan dan Kegiatan Sosial"),
    ("RSTU", "Jasa Lainnya"),
]

N_PROV = 34
N_SECT = 17
N_PS = N_PROV * N_SECT  # 578

REGION_ORDER = [
    "Sumatera",
    "Jawa",
    "Bali-Nusa Tenggara",
    "Kalimantan",
    "Sulawesi",
    "Maluku-Papua",
]


def clean_province_name(raw: str) -> str:
    """Remove leading province code like '11. ' from province name."""
    if raw is None:
        return ""
    s = str(raw).strip()
    for code, name, _ in PROVINCE_ORDER:
        if s == f"{code}. {name}" or s == name:
            return name
    # Fallback: strip leading "XX. "
    if len(s) > 3 and s[2] == "." and s[:2].isdigit():
        return s[4:].strip()
    if len(s) > 4 and s[3] == "." and s[:3].isdigit():
        return s[4:].strip()
    return s


def province_name_to_code(name: str) -> str:
    """Map cleaned province name to its code."""
    for code, pname, _ in PROVINCE_ORDER:
        if pname == name:
            return code
    raise ValueError(f"Unknown province: {name}")


def sector_code_to_name(code: str) -> str:
    for sc, sn in SECTOR_ORDER:
        if sc == code:
            return sn
    raise ValueError(f"Unknown sector code: {code}")


def extract_leontief_inverse(wb) -> np.ndarray:
    """Extract 578x578 Leontief inverse matrix from sheet."""
    ws = wb["Leontief_Inverse"]
    log.info("Extracting Leontief inverse matrix (578x578)...")

    # Validate labels match expected order
    expected_labels = []
    for _, pname, _ in PROVINCE_ORDER:
        for scode, _ in SECTOR_ORDER:
            expected_labels.append(f"{pname}|{scode}")

    # Row labels in col B, rows 2:579
    row_labels = []
    for r in range(2, 2 + N_PS):
        val = ws.cell(row=r, column=2).value
        row_labels.append(str(val).strip() if val else "")

    # Col labels in row 4, cols 5:582
    col_labels = []
    for c in range(5, 5 + N_PS):
        val = ws.cell(row=4, column=c).value
        col_labels.append(str(val).strip() if val else "")

    mismatches = 0
    for i, (exp, got) in enumerate(zip(expected_labels, row_labels)):
        if exp != got:
            log.warning(f"Row label mismatch at index {i}: expected '{exp}', got '{got}'")
            mismatches += 1
            if mismatches > 5:
                break

    for i, (exp, got) in enumerate(zip(expected_labels, col_labels)):
        if exp != got:
            log.warning(f"Col label mismatch at index {i}: expected '{exp}', got '{got}'")
            mismatches += 1
            if mismatches > 5:
                break

    if mismatches > 0:
        log.warning(f"Total label mismatches: {mismatches}. Proceeding with positional extraction.")

    # Extract matrix data: rows 5:582, cols 5:582
    matrix = np.zeros((N_PS, N_PS), dtype=np.float64)
    for i in range(N_PS):
        for j in range(N_PS):
            val = ws.cell(row=5 + i, column=5 + j).value
            if val is not None:
                matrix[i, j] = float(val)
            else:
                log.warning(f"None value in Leontief at ({i},{j})")

    # Validation: diagonal should be >= 1
    diag = np.diag(matrix)
    if np.any(diag < 1.0):
        bad = np.where(diag < 1.0)[0]
        log.warning(f"Leontief diagonal < 1 at indices: {bad[:10]}")
    else:
        log.info("Leontief diagonal validation PASS (all >= 1)")

    log.info(f"Leontief matrix shape: {matrix.shape}, min: {matrix.min():.6e}, max: {matrix.max():.6f}")
    return matrix


def extract_output_va_pdrb(wb) -> list[dict]:
    """Extract province-sector output, VA, VA coefficient data."""
    ws = wb["Output_VA_PDRB"]
    log.info("Extracting Output/VA/PDRB data...")

    records = []
    for r in range(2, 2 + N_PS):
        row_num = ws.cell(row=r, column=1).value
        prov_raw = ws.cell(row=r, column=2).value
        prov_code = ws.cell(row=r, column=3).value
        sector_name = ws.cell(row=r, column=4).value
        sector_code = ws.cell(row=r, column=5).value
        total_output = ws.cell(row=r, column=6).value  # Juta Rp
        va = ws.cell(row=r, column=7).value  # Juta Rp
        va_coeff = ws.cell(row=r, column=8).value

        prov_name = clean_province_name(prov_raw)

        if total_output is None or va is None:
            log.warning(f"Missing data at row {r}: output={total_output}, va={va}")
            continue

        # Compute VA coefficient if not available
        if va_coeff is None and total_output > 0:
            va_coeff = va / total_output

        records.append({
            "province_code": str(prov_code).strip(),
            "province_name": prov_name,
            "sector_code": str(sector_code).strip(),
            "sector_name": str(sector_name).strip(),
            "total_output_juta_rp": float(total_output),
            "value_added_juta_rp": float(va),
            "va_coefficient": float(va_coeff) if va_coeff else 0.0,
        })

    log.info(f"Extracted {len(records)} province-sector records")
    if len(records) != N_PS:
        log.error(f"Expected {N_PS} records, got {len(records)}")

    return records


def extract_province_pdrb(wb) -> list[dict]:
    """Extract PDRB summary per province from the summary section."""
    ws = wb["Output_VA_PDRB"]
    log.info("Extracting province PDRB summary...")

    pdrb_records = []
    # Summary section starts at row 582 (after 578 data rows + header + blank)
    for r in range(583, 583 + N_PROV):
        prov_raw = ws.cell(row=r, column=2).value
        total_output = ws.cell(row=r, column=3).value  # Juta Rp
        pdrb = ws.cell(row=r, column=4).value  # Juta Rp

        if prov_raw is None:
            continue

        prov_name = clean_province_name(prov_raw)
        prov_code = province_name_to_code(prov_name)

        pdrb_records.append({
            "province_code": prov_code,
            "province_name": prov_name,
            "total_output_juta_rp": float(total_output),
            "pdrb_juta_rp": float(pdrb),
            "total_output_miliar_rp": float(total_output) / 1000,
            "pdrb_miliar_rp": float(pdrb) / 1000,
        })

    log.info(f"Extracted {len(pdrb_records)} province PDRB records")
    return pdrb_records


def extract_bilateral_linkage(wb) -> tuple[np.ndarray, list[str]]:
    """Extract 34x34 bilateral linkage matrix from FL_BL sheet."""
    ws = wb["FL_BL"]
    log.info("Extracting bilateral linkage matrix...")

    # Bilateral section starts at row 587 (header row), data at row 588
    header_row = None
    for r in range(580, 660):
        val = ws.cell(row=r, column=1).value
        if val and "Provinsi Asal" in str(val):
            header_row = r
            break

    if header_row is None:
        log.error("Could not find bilateral linkage header row")
        return np.zeros((N_PROV, N_PROV)), []

    log.info(f"Bilateral linkage header at row {header_row}")

    # Column labels (provinces)
    col_labels = []
    for c in range(2, 2 + N_PROV):
        val = ws.cell(row=header_row, column=c).value
        col_labels.append(clean_province_name(val) if val else "")

    # Extract matrix
    matrix = np.zeros((N_PROV, N_PROV), dtype=np.float64)
    row_labels = []
    for i in range(N_PROV):
        r = header_row + 1 + i
        rl = ws.cell(row=r, column=1).value
        row_labels.append(clean_province_name(rl) if rl else "")
        for j in range(N_PROV):
            val = ws.cell(row=r, column=2 + j).value
            if val is not None:
                matrix[i, j] = float(val)

    log.info(f"Bilateral matrix shape: {matrix.shape}")
    return matrix, row_labels


def extract_fl_bl_578(wb) -> list[dict]:
    """Extract FL/BL indices for all 578 province-sector combinations."""
    ws = wb["FL_BL"]
    log.info("Extracting FL/BL data for 578 province-sector combinations...")

    records = []
    for r in range(3, 3 + N_PS):
        row_num = ws.cell(row=r, column=1).value
        prov_raw = ws.cell(row=r, column=2).value
        sector_name = ws.cell(row=r, column=3).value
        sector_code = ws.cell(row=r, column=4).value
        bl_raw = ws.cell(row=r, column=5).value
        fl_raw = ws.cell(row=r, column=6).value
        bl_index = ws.cell(row=r, column=7).value
        fl_index = ws.cell(row=r, column=8).value
        classification = ws.cell(row=r, column=9).value

        prov_name = clean_province_name(prov_raw)

        records.append({
            "province_name": prov_name,
            "province_code": province_name_to_code(prov_name) if prov_name else "",
            "sector_code": str(sector_code).strip() if sector_code else "",
            "sector_name": str(sector_name).strip() if sector_name else "",
            "bl_raw": float(bl_raw) if bl_raw is not None else None,
            "fl_raw": float(fl_raw) if fl_raw is not None else None,
            "bl_index": float(bl_index) if bl_index is not None else None,
            "fl_index": float(fl_index) if fl_index is not None else None,
            "classification": str(classification) if classification else None,
        })

    log.info(f"Extracted {len(records)} FL/BL records")
    return records


def compute_fl_bl_from_leontief(L: np.ndarray) -> list[dict]:
    """Compute FL/BL indices directly from the Leontief inverse matrix.

    BL_j = (1/n) * SUM_i(L_ij) / [(1/n^2) * SUM_all(L)]
    FL_i = (1/n) * SUM_j(L_ij) / [(1/n^2) * SUM_all(L)]
    """
    n = L.shape[0]  # 578
    total_sum = L.sum()
    global_avg = total_sum / (n * n)

    col_sums = L.sum(axis=0)  # BL raw
    row_sums = L.sum(axis=1)  # FL raw

    bl_index = (col_sums / n) / global_avg
    fl_index = (row_sums / n) / global_avg

    records = []
    idx = 0
    for pcode, pname, region in PROVINCE_ORDER:
        for scode, sname in SECTOR_ORDER:
            classification = _classify_linkage(bl_index[idx], fl_index[idx])
            records.append({
                "province_code": pcode,
                "province_name": pname,
                "sector_code": scode,
                "sector_name": sname,
                "bl_raw": float(col_sums[idx]),
                "fl_raw": float(row_sums[idx]),
                "bl_index": float(bl_index[idx]),
                "fl_index": float(fl_index[idx]),
                "classification": classification,
            })
            idx += 1

    return records


def _classify_linkage(bl: float, fl: float) -> str:
    if bl > 1 and fl > 1:
        return "Key Sector (BL>1, FL>1)"
    elif bl > 1 and fl <= 1:
        return "Strong BL (BL>1, FL<1)"
    elif bl <= 1 and fl > 1:
        return "Strong FL (BL<1, FL>1)"
    else:
        return "Weak (BL<1, FL<1)"


def compute_bilateral_from_leontief(L: np.ndarray) -> np.ndarray:
    """Compute 34x34 bilateral linkage matrix by aggregating 17x17 sub-blocks.

    bilateral[i,j] = SUM over all sector pairs of L[sectors_of_prov_i, sectors_of_prov_j]
    """
    bilateral = np.zeros((N_PROV, N_PROV), dtype=np.float64)
    for i in range(N_PROV):
        for j in range(N_PROV):
            row_start = i * N_SECT
            row_end = row_start + N_SECT
            col_start = j * N_SECT
            col_end = col_start + N_SECT
            bilateral[i, j] = L[row_start:row_end, col_start:col_end].sum()
    return bilateral


def build_province_metadata() -> list[dict]:
    return [
        {
            "id": i,
            "code": code,
            "name": name,
            "region": region,
        }
        for i, (code, name, region) in enumerate(PROVINCE_ORDER)
    ]


def build_sector_metadata() -> list[dict]:
    return [
        {
            "id": i,
            "code": code,
            "name": name,
        }
        for i, (code, name) in enumerate(SECTOR_ORDER)
    ]


def build_region_mapping() -> dict:
    regions = {}
    for code, name, region in PROVINCE_ORDER:
        if region not in regions:
            regions[region] = []
        regions[region].append({"code": code, "name": name})
    return regions


def validate_data(
    leontief: np.ndarray,
    ps_data: list[dict],
    pdrb_data: list[dict],
    fl_bl: list[dict],
    bilateral: np.ndarray,
):
    """Run validation checks analogous to the workbook's Validation_Check sheet."""
    log.info("=" * 60)
    log.info("VALIDATION")
    log.info("=" * 60)

    errors = 0

    # CHECK: Province-sector count
    if len(ps_data) == N_PS:
        log.info(f"CHECK: Province-sector count = {N_PS} ... PASS")
    else:
        log.error(f"CHECK: Province-sector count = {len(ps_data)}, expected {N_PS} ... FAIL")
        errors += 1

    # CHECK: Province count
    if len(pdrb_data) == N_PROV:
        log.info(f"CHECK: Province count = {N_PROV} ... PASS")
    else:
        log.error(f"CHECK: Province count = {len(pdrb_data)}, expected {N_PROV} ... FAIL")
        errors += 1

    # CHECK: Leontief dimensions
    if leontief.shape == (N_PS, N_PS):
        log.info(f"CHECK: Leontief matrix shape = {leontief.shape} ... PASS")
    else:
        log.error(f"CHECK: Leontief matrix shape = {leontief.shape}, expected ({N_PS},{N_PS}) ... FAIL")
        errors += 1

    # CHECK: Leontief diagonal >= 1
    diag = np.diag(leontief)
    if np.all(diag >= 1.0):
        log.info(f"CHECK: Leontief diagonal all >= 1 ... PASS (min: {diag.min():.6f})")
    else:
        bad_count = np.sum(diag < 1.0)
        log.error(f"CHECK: Leontief diagonal has {bad_count} values < 1 ... FAIL")
        errors += 1

    # CHECK: VA coefficients in [0, 1]
    va_coeffs = [r["va_coefficient"] for r in ps_data]
    out_of_range = sum(1 for v in va_coeffs if v < 0 or v > 1)
    if out_of_range == 0:
        log.info(f"CHECK: VA coefficients all in [0,1] ... PASS")
    else:
        log.warning(f"CHECK: {out_of_range} VA coefficients outside [0,1] ... WARNING")

    # CHECK: All outputs > 0
    zero_output = sum(1 for r in ps_data if r["total_output_juta_rp"] <= 0)
    if zero_output == 0:
        log.info(f"CHECK: All sector outputs > 0 ... PASS")
    else:
        log.warning(f"CHECK: {zero_output} sectors with output <= 0 ... WARNING")

    # CHECK: Unique province-sector keys
    keys = set()
    for r in ps_data:
        k = f"{r['province_code']}|{r['sector_code']}"
        if k in keys:
            log.error(f"CHECK: Duplicate province-sector key: {k} ... FAIL")
            errors += 1
        keys.add(k)
    if len(keys) == N_PS:
        log.info(f"CHECK: Unique province-sector keys = {N_PS} ... PASS")

    # CHECK: Bilateral matrix dimensions
    if bilateral.shape == (N_PROV, N_PROV):
        log.info(f"CHECK: Bilateral matrix shape = {bilateral.shape} ... PASS")
    else:
        log.error(f"CHECK: Bilateral matrix shape ... FAIL")
        errors += 1

    # CHECK: PDRB consistency (sum of VA per province ≈ PDRB)
    for prov in pdrb_data:
        pcode = prov["province_code"]
        va_sum = sum(r["value_added_juta_rp"] for r in ps_data if r["province_code"] == pcode)
        pdrb = prov["pdrb_juta_rp"]
        if pdrb > 0:
            rel_err = abs(va_sum - pdrb) / pdrb
            if rel_err > 0.001:
                log.warning(f"CHECK: PDRB consistency for {prov['province_name']}: VA sum={va_sum:.0f}, PDRB={pdrb:.0f}, rel_err={rel_err:.6f}")

    log.info(f"Validation complete: {errors} errors found")
    return errors


def main():
    if not WORKBOOK_PATH.exists():
        log.error(f"Workbook not found at {WORKBOOK_PATH}")
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    log.info(f"Loading workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(str(WORKBOOK_PATH), data_only=True)
    log.info(f"Sheets: {wb.sheetnames}")

    # 1. Extract Leontief inverse
    leontief = extract_leontief_inverse(wb)
    np.save(OUT_DIR / "leontief_inverse.npy", leontief)
    log.info(f"Saved leontief_inverse.npy ({leontief.nbytes / 1e6:.1f} MB)")

    # 2. Extract Output/VA/PDRB per province-sector
    ps_data = extract_output_va_pdrb(wb)
    with open(OUT_DIR / "province_sector_data.json", "w", encoding="utf-8") as f:
        json.dump(ps_data, f, ensure_ascii=False, indent=2)

    # 3. Extract province PDRB summary
    pdrb_data = extract_province_pdrb(wb)
    with open(OUT_DIR / "province_pdrb.json", "w", encoding="utf-8") as f:
        json.dump(pdrb_data, f, ensure_ascii=False, indent=2)

    # 4. Compute FL/BL from Leontief (more reliable than reading formula cells)
    fl_bl = compute_fl_bl_from_leontief(leontief)
    with open(OUT_DIR / "fl_bl_data.json", "w", encoding="utf-8") as f:
        json.dump(fl_bl, f, ensure_ascii=False, indent=2)

    # 5. Compute bilateral linkage from Leontief
    bilateral = compute_bilateral_from_leontief(leontief)
    np.save(OUT_DIR / "bilateral_linkage.npy", bilateral)

    bilateral_labels = [name for _, name, _ in PROVINCE_ORDER]
    with open(OUT_DIR / "bilateral_labels.json", "w", encoding="utf-8") as f:
        json.dump(bilateral_labels, f, ensure_ascii=False, indent=2)

    # 6. Province metadata
    prov_meta = build_province_metadata()
    with open(OUT_DIR / "province_metadata.json", "w", encoding="utf-8") as f:
        json.dump(prov_meta, f, ensure_ascii=False, indent=2)

    # 7. Sector metadata
    sect_meta = build_sector_metadata()
    with open(OUT_DIR / "sector_metadata.json", "w", encoding="utf-8") as f:
        json.dump(sect_meta, f, ensure_ascii=False, indent=2)

    # 8. Region mapping
    region_map = build_region_mapping()
    with open(OUT_DIR / "region_mapping.json", "w", encoding="utf-8") as f:
        json.dump(region_map, f, ensure_ascii=False, indent=2)

    # 9. VA coefficient vector (aligned with Leontief matrix rows)
    va_coefficients = np.array([r["va_coefficient"] for r in ps_data], dtype=np.float64)
    np.save(OUT_DIR / "va_coefficients.npy", va_coefficients)

    # 10. Output vector
    output_vector = np.array([r["total_output_juta_rp"] for r in ps_data], dtype=np.float64)
    np.save(OUT_DIR / "output_vector.npy", output_vector)

    # Validation
    validate_data(leontief, ps_data, pdrb_data, fl_bl, bilateral)

    log.info("=" * 60)
    log.info("ETL COMPLETE")
    log.info(f"Output directory: {OUT_DIR}")
    log.info(f"Files: {[f.name for f in sorted(OUT_DIR.iterdir())]}")


if __name__ == "__main__":
    main()
